#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Figure 1e: transition-pathway concentration in UP hexagons.

Refined v4:
- same fixed canvas and main-axis geometry as Fig. 1d refined v2;
- rank/pathway labels are drawn on a separate label axis;
- percentage labels are smaller and placed outside the bars;
- cumulative line and markers are restrained;
- fixed-canvas export, no bbox_inches='tight';
- legend moved upward for a tighter panel composition.
"""

from pathlib import Path
from openpyxl import load_workbook
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
from matplotlib.lines import Line2D

# =============================================================================
# User-defined paths
# =============================================================================
INPUT_XLSX = Path(r"Landcover_change/change matrix up area.xlsx")
OUTPUT_PNG = Path(r"Fig1e_pathway_concentration_horizontal_refined_v8_leftspine.png")
OUTPUT_PDF = Path(r"Fig1e_pathway_concentration_horizontal_refined_v8_leftspine.pdf")
OUTPUT_SVG = Path(r"Fig1e_pathway_concentration_horizontal_refined_v8_leftspine.svg")
OUTPUT_CSV = Path(r"Fig1e_pathway_concentration_horizontal_refined_v8_leftspine.csv")

EXPORT_TRANSPARENT = True

# =============================================================================
# Panel geometry: same main-axis position as Fig. 1d refined v2
# =============================================================================
FIGSIZE = (3.80, 2.55)
DPI = 600
FONT_FAMILY = "Arial"
AX_POS = [0.36, 0.255, 0.59, 0.625]
LABEL_AX_POS = [0.035, AX_POS[1], 0.305, AX_POS[3]]
# Legend remains aligned to the main plotting axis but is slightly compacted
# so the rightmost item stays inside the fixed canvas.
LEGEND_X = AX_POS[0] + AX_POS[2] / 2 - 0.010
LEGEND_Y = 0.095

# =============================================================================
# Styling
# =============================================================================
BASE_FONT = 6.8
LABEL_FONT = 6.45
TICK_FONT = 6.35
AXIS_LABEL_FONT = 7.05
PERCENT_FONT = 5.70
NOTE_FONT = 5.85
LEGEND_FONT = 5.45

BAR_HEIGHT = 0.48
CUM_LINEWIDTH = 0.92
CUM_MARKERSIZE = 2.35
GRID_LINEWIDTH = 0.42
AXIS_LINEWIDTH = 0.62

ARROW = "\u2192"

CLASS_NAMES = [
    "Cropland", "Forest", "Shrubland", "Grassland", "Lichen & moss",
    "Sparse vegetation", "Wetland", "Urban", "Bare areas", "Water & ice"
]
LOWER_COVER_CLASSES = {"Bare areas", "Sparse vegetation"}
HIGHER_COVER_NATURAL = {"Forest", "Grassland", "Lichen & moss", "Shrubland", "Wetland"}

# Keep these semantic colours harmonized with Fig. 1d and Fig. 1c.
GROUP_COLORS = {
    "Greening / recovery": "#77B98A",
    "Degradation / browning": "#D7A257",
    "Agricultural expansion": "#9ECAE1",
    "Urban expansion": "#8D8D8D",
    "Other transition": "#B5B5B5",
}
LINE_COLOR = "#333333"
GRID_COLOR = "#E7E7E7"
AXIS_COLOR = "#333333"
TEXT_COLOR = "#222222"

SHOW_TITLE = False


def read_matrix(path: Path) -> np.ndarray:
    """Read the 10 x 10 land-cover transition matrix from the first worksheet."""
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    matrix = np.zeros((10, 10), dtype=float)
    for r in range(10):
        for c in range(10):
            value = ws.cell(r + 2, c + 2).value
            matrix[r, c] = float(value) if value is not None else 0.0
    return matrix


def classify_transition(fr: str, to: str) -> str:
    """Assign each off-diagonal transition to a semantic group for plotting."""
    if to == "Urban":
        return "Urban expansion"
    if to == "Cropland" and fr != "Cropland":
        return "Agricultural expansion"
    if fr in LOWER_COVER_CLASSES and to in HIGHER_COVER_NATURAL:
        return "Greening / recovery"
    if fr == "Bare areas" and to == "Sparse vegetation":
        return "Greening / recovery"
    if (
        (fr == "Grassland" and to in {"Sparse vegetation", "Bare areas"})
        or (fr == "Sparse vegetation" and to == "Bare areas")
        or (fr in HIGHER_COVER_NATURAL and to == "Sparse vegetation")
    ):
        return "Degradation / browning"
    return "Other transition"


def short_class_name(name: str) -> str:
    replacements = {
        "Bare areas": "Bare",
        "Sparse vegetation": "Sparse veg.",
        "Grassland": "Grass",
        "Cropland": "Crop",
        "Shrubland": "Shrub",
        "Lichen & moss": "Lichen/moss",
        "Water & ice": "Water/ice",
    }
    return replacements.get(name, name)


def short_pathway_label(fr: str, to: str) -> str:
    return f"{short_class_name(fr)} {ARROW} {short_class_name(to)}"


def main() -> None:
    plt.rcParams.update({
        "font.family": FONT_FAMILY,
        "font.sans-serif": [FONT_FAMILY],
        "font.size": BASE_FONT,
        "axes.linewidth": AXIS_LINEWIDTH,
        "pdf.fonttype": 42,
        "ps.fonttype": 42,
        "svg.fonttype": "none",
        "xtick.major.size": 2.4,
        "ytick.major.size": 0,
        "xtick.major.width": AXIS_LINEWIDTH,
        "ytick.major.width": 0,
        "axes.unicode_minus": False,
    })

    matrix = read_matrix(INPUT_XLSX)

    rows = []
    for i in range(10):
        for j in range(10):
            if i == j:
                continue
            area = float(matrix[i, j])
            if area <= 0:
                continue
            fr = CLASS_NAMES[i]
            to = CLASS_NAMES[j]
            group = classify_transition(fr, to)
            rows.append({
                "from_id": i + 1,
                "to_id": j + 1,
                "from_class": fr,
                "to_class": to,
                "pathway": f"{fr} -> {to}",
                "pathway_short": short_pathway_label(fr, to),
                "area": area,
                "semantic_group": group,
                "color": GROUP_COLORS[group],
            })

    df_all = pd.DataFrame(rows).sort_values("area", ascending=False).reset_index(drop=True)
    df_all["rank"] = np.arange(1, len(df_all) + 1)
    offdiag_total = float(df_all["area"].sum())
    df_all["share_pct"] = df_all["area"] / offdiag_total * 100.0
    df_all["cum_share_pct"] = df_all["share_pct"].cumsum()

    plot = df_all.head(10).copy()
    plot.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")

    top3 = float(plot.loc[plot["rank"] <= 3, "share_pct"].sum())
    top10 = float(plot["share_pct"].sum())

    fig = plt.figure(figsize=FIGSIZE, dpi=DPI, facecolor="white")
    ax_label = fig.add_axes(LABEL_AX_POS)
    ax = fig.add_axes(AX_POS)

    y = np.arange(len(plot))
    bar_vals = plot["share_pct"].values.astype(float)
    cum_vals = plot["cum_share_pct"].values.astype(float)
    colors = plot["color"].values

    bars = ax.barh(
        y,
        bar_vals,
        color=colors,
        edgecolor="none",
        height=BAR_HEIGHT,
        zorder=3,
    )

    ax.plot(
        cum_vals,
        y,
        color=LINE_COLOR,
        linewidth=CUM_LINEWIDTH,
        marker="o",
        markersize=CUM_MARKERSIZE,
        markerfacecolor=LINE_COLOR,
        markeredgecolor=LINE_COLOR,
        zorder=5,
    )

    ax.invert_yaxis()
    ax.set_xlim(0, 100)
    ax.set_ylim(len(plot) - 0.35, -0.75)
    ax.set_yticks(y)
    ax.set_yticklabels([])
    ax.set_xticks([0, 25, 50, 75, 100])
    ax.set_xticklabels(["0", "25", "50", "75", "100"], fontsize=TICK_FONT, color=AXIS_COLOR)
    ax.set_xlabel(
        "Share of off-diagonal transition area (%)",
        fontsize=AXIS_LABEL_FONT,
        color=TEXT_COLOR,
        labelpad=4,
    )

    if SHOW_TITLE:
        ax.set_title("Transition-pathway concentration", fontsize=7.2, color=TEXT_COLOR, pad=4)

    ax.grid(axis="x", color=GRID_COLOR, linewidth=GRID_LINEWIDTH, zorder=1)
    ax.grid(axis="y", visible=False)
    ax.set_axisbelow(True)

    # Percentage labels for the leading three pathways only.
    # All labels are placed outside the bars, as requested.
    for idx, (bar, value) in enumerate(zip(bars, bar_vals)):
        if idx > 2:
            continue
        ymid = bar.get_y() + bar.get_height() / 2
        # Move the first label a little farther right to avoid the cumulative dot.
        x_pad = 2.2 if idx == 0 else 1.25
        ax.text(
            value + x_pad,
            ymid,
            f"{value:.1f}%",
            va="center",
            ha="left",
            fontsize=PERCENT_FONT,
            color=TEXT_COLOR,
            zorder=6,
        )

    # Summary note placed in the upper-right blank area, away from the cumulative line.
    ax.text(
        0.98,
        0.98,
        f"Top 3 = {top3:.1f}%\nTop 10 = {top10:.1f}%",
        transform=ax.transAxes,
        ha="right",
        va="top",
        fontsize=NOTE_FONT,
        color=TEXT_COLOR,
        linespacing=1.05,
        zorder=6,
    )

    # Keep left + bottom borders to match the rest of Fig. 1.  The separate
    # label axis remains borderless; the visible left spine belongs to the
    # actual plotting region.
    for spine in ["top", "right"]:
        ax.spines[spine].set_visible(False)
    ax.spines["left"].set_visible(True)
    ax.spines["left"].set_color(AXIS_COLOR)
    ax.spines["left"].set_linewidth(AXIS_LINEWIDTH)
    ax.spines["bottom"].set_color(AXIS_COLOR)
    ax.spines["bottom"].set_linewidth(AXIS_LINEWIDTH)
    ax.tick_params(axis="x", colors=AXIS_COLOR, labelsize=TICK_FONT, length=2.4, pad=2)
    ax.tick_params(axis="y", length=0)

    # Label axis: rank and pathway labels in two aligned columns.
    ax_label.set_xlim(0, 1)
    ax_label.set_ylim(ax.get_ylim())
    ax_label.axis("off")
    for yi, rank, label in zip(y, plot["rank"].values, plot["pathway_short"].values):
        ax_label.text(
            0.12,
            yi,
            f"{int(rank)}",
            ha="right",
            va="center",
            fontsize=LABEL_FONT,
            color=TEXT_COLOR,
            clip_on=False,
        )
        ax_label.text(
            0.20,
            yi,
            label,
            ha="left",
            va="center",
            fontsize=LABEL_FONT,
            color=TEXT_COLOR,
            clip_on=False,
        )

    legend_handles = [
        Patch(facecolor=GROUP_COLORS["Greening / recovery"], edgecolor="none", label="Greening"),
        Patch(facecolor=GROUP_COLORS["Degradation / browning"], edgecolor="none", label="Browning"),
        Patch(facecolor=GROUP_COLORS["Agricultural expansion"], edgecolor="none", label="Agriculture"),
        Patch(facecolor=GROUP_COLORS["Urban expansion"], edgecolor="none", label="Urban"),
        Line2D(
            [0], [0],
            color=LINE_COLOR,
            marker="o",
            linewidth=CUM_LINEWIDTH,
            markersize=CUM_MARKERSIZE,
            label="Cumulative",
        ),
    ]
    # Legend is centred under the main plotting axis, not the full canvas.
    # This aligns it with the x-axis title and the data region.
    fig.legend(
        handles=legend_handles,
        loc="lower center",
        bbox_to_anchor=(LEGEND_X, LEGEND_Y),
        ncol=5,
        frameon=False,
        fontsize=LEGEND_FONT,
        handlelength=0.82,
        handletextpad=0.28,
        columnspacing=0.36,
        borderaxespad=0,
    )

    # Fixed-canvas export for alignment. Do not use bbox_inches='tight'.
    if EXPORT_TRANSPARENT:
        fig.patch.set_alpha(0)
        for a in fig.axes:
            a.set_facecolor("none")

    save_kwargs = dict(transparent=EXPORT_TRANSPARENT)
    fig.savefig(OUTPUT_PNG, dpi=DPI, **save_kwargs)
    fig.savefig(OUTPUT_PDF, **save_kwargs)
    fig.savefig(OUTPUT_SVG, **save_kwargs)
    plt.close(fig)

    print(f"Saved: {OUTPUT_PNG}")
    print(f"Saved: {OUTPUT_PDF}")
    print(f"Saved: {OUTPUT_SVG}")
    print(f"Saved: {OUTPUT_CSV}")


if __name__ == "__main__":
    main()
