from openpyxl import load_workbook
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
plt.rcParams["font.family"] = "Arial"
plt.rcParams["font.sans-serif"] = ["Arial"]
from matplotlib.path import Path
from matplotlib.patches import PathPatch, Rectangle

infile = r"Landcover_change/change matrix up area.xlsx"
out_png = r"Fig1_UP_alluvial_top_flows.png"
out_pdf = r"Fig1_UP_alluvial_top_flows.pdf"
out_csv = r"Fig1_UP_alluvial_top_flows.csv"

class_names = ['Cropland', 'Forest', 'Shrubland', 'Grassland', 'Lichen & moss', 'Sparse vegetation', 'Wetland', 'Urban', 'Bare areas', 'Water & ice']

wb = load_workbook(infile, data_only=True)
ws = wb[wb.sheetnames[0]]
matrix = np.zeros((10, 10), dtype=float)
for r in range(10):
    for c in range(10):
        v = ws.cell(r + 2, c + 2).value
        matrix[r, c] = float(v) if v is not None else 0.0

rows = []
for i in range(10):
    for j in range(10):
        area = float(matrix[i, j])
        rows.append({
            "from_id": i + 1,
            "to_id": j + 1,
            "from_class": class_names[i],
            "to_class": class_names[j],
            "area": area,
            "is_diagonal": i == j
        })
df = pd.DataFrame(rows)

offdiag = df[(~df["is_diagonal"]) & (df["area"] > 0)].sort_values("area", ascending=False).copy()
top_offdiag_n = 10
top_offdiag = offdiag.head(top_offdiag_n).copy()
diag = df[df["is_diagonal"]].copy()
plot_df = pd.concat([diag, top_offdiag], ignore_index=True)
plot_df.to_csv(out_csv, index=False, encoding="utf-8-sig")

left_totals = plot_df.groupby("from_class", as_index=False)["area"].sum().rename(columns={"area": "left_total"})
right_totals = plot_df.groupby("to_class", as_index=False)["area"].sum().rename(columns={"area": "right_total"})
left_totals = left_totals.set_index("from_class").reindex(class_names).fillna(0).reset_index()
right_totals = right_totals.set_index("to_class").reindex(class_names).fillna(0).reset_index()

fig, ax = plt.subplots(figsize=(11.5, 8.2), dpi=220)
ax.set_xlim(0, 1)
ax.set_ylim(0, 1)
ax.axis("off")

left_x0, left_x1 = 0.08, 0.16
right_x0, right_x1 = 0.84, 0.92
top_margin = 0.04
bottom_margin = 0.04
usable_h = 1 - top_margin - bottom_margin
gap = 0.008

left_sum = left_totals["left_total"].sum()
right_sum = right_totals["right_total"].sum()
left_heights = [0 if left_sum == 0 else usable_h * (v / left_sum) for v in left_totals["left_total"]]
right_heights = [0 if right_sum == 0 else usable_h * (v / right_sum) for v in right_totals["right_total"]]

def positions_from_heights(heights):
    nonzero = sum(h > 0 for h in heights)
    gap_total = gap * max(nonzero - 1, 0)
    scale = (usable_h - gap_total) / sum(heights) if sum(heights) > 0 else 1
    scaled = [h * scale for h in heights]
    pos = {}
    y = 1 - top_margin
    for name, h in zip(class_names, scaled):
        if h <= 0:
            pos[name] = (None, None)
            continue
        y1 = y
        y0 = y - h
        pos[name] = (y0, y1)
        y = y0 - gap
    return pos, scaled

left_pos, left_scaled = positions_from_heights(left_heights)
right_pos, right_scaled = positions_from_heights(right_heights)

higher_cover_natural = ['Forest', 'Grassland', 'Lichen & moss', 'Shrubland', 'Wetland']
lower_cover_classes = ['Bare areas', 'Sparse vegetation']

def flow_color(fr, to, diag=False):
    if diag:
        return "#D9D9D9"
    if to == "Urban":
        return "#9A9A9A"
    if to == "Cropland" and fr != "Cropland":
        return "#7FA3CC"
    if (fr in lower_cover_classes and to in higher_cover_natural) or (fr == "Bare areas" and to == "Sparse vegetation"):
        return "#7CC08A"
    if (fr == "Grassland" and to in {"Sparse vegetation", "Bare areas"}) or        (fr == "Sparse vegetation" and to == "Bare areas") or        (fr in higher_cover_natural and to == "Sparse vegetation"):
        return "#D8A15A"
    return "#B6B6B6"

bar_colors = {
    "Cropland": "#A7C4E4",
    "Forest": "#B9D9B3",
    "Shrubland": "#C7DFC1",
    "Grassland": "#9FD090",
    "Lichen & moss": "#DCE8C8",
    "Sparse vegetation": "#B7D9A9",
    "Wetland": "#B8DCCF",
    "Urban": "#C4C4C4",
    "Bare areas": "#E8D9B5",
    "Water & ice": "#D7E6EF",
}

for name in class_names:
    y0, y1 = left_pos[name]
    if y0 is None:
        continue
    ax.add_patch(Rectangle((left_x0, y0), left_x1-left_x0, y1-y0,
                           facecolor=bar_colors.get(name, "#DDDDDD"), edgecolor="white", linewidth=1.0))
    ax.text(left_x0 - 0.012, (y0+y1)/2, name, ha="right", va="center", fontsize=9.2, color="#222222")

for name in class_names:
    y0, y1 = right_pos[name]
    if y0 is None:
        continue
    ax.add_patch(Rectangle((right_x0, y0), right_x1-right_x0, y1-y0,
                           facecolor=bar_colors.get(name, "#DDDDDD"), edgecolor="white", linewidth=1.0))
    ax.text(right_x1 + 0.012, (y0+y1)/2, name, ha="left", va="center", fontsize=9.2, color="#222222")

ax.text((left_x0+left_x1)/2, 1-top_margin/2, "2000", ha="center", va="center", fontsize=11, color="#222222")
ax.text((right_x0+right_x1)/2, 1-top_margin/2, "2022", ha="center", va="center", fontsize=11, color="#222222")
ax.set_title("Major land-cover flows in greening hexagons", fontsize=12, color="#222222", pad=10)

left_offsets = {name: left_pos[name][1] if left_pos[name][1] is not None else None for name in class_names}
right_offsets = {name: right_pos[name][1] if right_pos[name][1] is not None else None for name in class_names}

left_scale = (usable_h - gap * max(sum(h > 0 for h in left_heights)-1, 0)) / left_sum if left_sum > 0 else 1
right_scale = (usable_h - gap * max(sum(h > 0 for h in right_heights)-1, 0)) / right_sum if right_sum > 0 else 1

plot_df = plot_df.sort_values(["from_class", "area"], ascending=[True, False]).reset_index(drop=True)

def ribbon(ax, x0, x1, y0b, y0t, y1b, y1t, color, alpha=0.72):
    c = 0.28
    verts = [
        (x0, y0t),
        (x0 + c*(x1-x0), y0t),
        (x1 - c*(x1-x0), y1t),
        (x1, y1t),
        (x1, y1b),
        (x1 - c*(x1-x0), y1b),
        (x0 + c*(x1-x0), y0b),
        (x0, y0b),
        (x0, y0t)
    ]
    codes = [
        Path.MOVETO,
        Path.CURVE4,
        Path.CURVE4,
        Path.CURVE4,
        Path.LINETO,
        Path.CURVE4,
        Path.CURVE4,
        Path.CURVE4,
        Path.CLOSEPOLY
    ]
    ax.add_patch(PathPatch(Path(verts, codes), facecolor=color, edgecolor="none", alpha=alpha))

for _, r in plot_df.iterrows():
    fr = r["from_class"]
    to = r["to_class"]
    area = float(r["area"])
    h_left = area * left_scale
    h_right = area * right_scale
    if left_offsets[fr] is None or right_offsets[to] is None:
        continue
    y0t = left_offsets[fr]
    y0b = y0t - h_left
    left_offsets[fr] = y0b
    y1t = right_offsets[to]
    y1b = y1t - h_right
    right_offsets[to] = y1b
    col = flow_color(fr, to, diag=(fr == to))
    alpha = 0.38 if fr == to else 0.78
    ribbon(ax, left_x1, right_x0, y0b, y0t, y1b, y1t, col, alpha=alpha)

legend_lines = [
    ("Greening / recovery", "#7CC08A"),
    ("Degradation / browning", "#D8A15A"),
    ("Agricultural expansion", "#7FA3CC"),
    ("Urban expansion", "#9A9A9A"),
    ("Within-class persistence", "#D9D9D9"),
]
lx, ly = 0.08, 0.015
for i, (lab, col) in enumerate(legend_lines):
    x = lx + i * 0.18
    ax.add_patch(Rectangle((x, ly), 0.015, 0.012, facecolor=col, edgecolor="none"))
    ax.text(x + 0.02, ly + 0.006, lab, va="center", ha="left", fontsize=8.3, color="#444444")

fig.tight_layout()
fig.savefig(out_png, bbox_inches="tight", transparent=True)
fig.savefig(out_pdf, bbox_inches="tight", transparent=True)
