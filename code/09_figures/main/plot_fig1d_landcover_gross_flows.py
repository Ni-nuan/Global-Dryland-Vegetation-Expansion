#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Figure 1d: gross inflow and outflow by aggregated land-cover class in UP hexagons.

Refined version for side-by-side assembly with Fig. 1e:
- mirrored horizontal bars retained;
- outflow colour matched to the Fig. 1b/c light-blue system (#9ECAE1);
- full-number tick labels and key value labels retained;
- all value labels remain outside bars;
- fixed canvas export, no bbox_inches='tight';
- legend placed below the chart to match Fig. 1e.
"""

from pathlib import Path
from openpyxl import load_workbook
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.patches import Patch

# =============================================================================
# User-defined paths
# =============================================================================
INPUT_XLSX = Path(r"Landcover_change/change matrix up area.xlsx")
OUTPUT_PNG = Path(r"Fig1d_gross_inflow_outflow_refined_v5_leftspine.png")
OUTPUT_PDF = Path(r"Fig1d_gross_inflow_outflow_refined_v5_leftspine.pdf")
OUTPUT_SVG = Path(r"Fig1d_gross_inflow_outflow_refined_v5_leftspine.svg")
OUTPUT_CSV = Path(r"Fig1d_gross_inflow_outflow_refined_v5_leftspine.csv")

EXPORT_TRANSPARENT = True

# =============================================================================
# Panel geometry: keep consistent with Fig. 1e refined panel
# =============================================================================
FIGSIZE = (3.80, 2.55)
DPI = 600
AX_POS = [0.36, 0.255, 0.59, 0.625]  # left, bottom, width, height
LEGEND_X = AX_POS[0] + AX_POS[2] / 2
LEGEND_Y = 0.095
FONT_FAMILY = "Arial"

# =============================================================================
# Styling
# =============================================================================
BASE_FONT = 6.8
LABEL_FONT = 6.45
TICK_FONT = 6.35
AXIS_LABEL_FONT = 7.05
LEGEND_FONT = 5.90
VALUE_FONT = 5.85

# Blue is matched to Fig. 1b/c IQR fill. Sand is matched to Fig. 1e browning.
OUTFLOW_COLOR = "#9ECAE1"
INFLOW_COLOR = "#D7A257"
GRID_COLOR = "#E7E7E7"
AXIS_COLOR = "#333333"
TEXT_COLOR = "#222222"
ZERO_COLOR = "#555555"

BAR_HEIGHT = 0.44
GRID_LINEWIDTH = 0.42
AXIS_LINEWIDTH = 0.62
ZERO_LINEWIDTH = 0.70

# Only label major values; keep minor categories readable through bar lengths.
LABEL_THRESHOLD = 30000.0
LABEL_PAD_FRAC = 0.018

# Extra horizontal space for outside labels. Increase LEFT_LABEL_ZONE if the
# longest outflow label approaches the class labels after assembly.
LEFT_LABEL_ZONE = 0.34
RIGHT_LABEL_ZONE = 0.28

CLASS_NAMES = [
    "Cropland", "Forest", "Shrubland", "Grassland", "Lichen & moss",
    "Sparse vegetation", "Wetland", "Urban", "Bare areas", "Water & ice"
]

DISPLAY_ORDER = [
    "Bare areas", "Sparse vegetation", "Grassland", "Cropland", "Shrubland",
    "Urban", "Water & ice", "Forest", "Wetland", "Lichen & moss"
]


def read_matrix(path: Path) -> np.ndarray:
    """Read the 10 x 10 land-cover transition matrix from the first worksheet."""
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    matrix = np.zeros((10, 10), dtype=float)
    for r in range(10):
        for c in range(10):
            value = ws.cell(r + 2, c + 2).value
            matrix[r, c] = float(value) if value is not None else 0.0
    return matrix


def fmt_full(v: float) -> str:
    return f"{v:,.0f}"


def main() -> None:
    plt.rcParams.update({
        "font.family": FONT_FAMILY,
        "font.sans-serif": [FONT_FAMILY],
        "font.size": BASE_FONT,
        "axes.linewidth": AXIS_LINEWIDTH,
        "pdf.fonttype": 42,
        "ps.fonttype": 42,
        "svg.fonttype": "none",
        "xtick.major.size": 2.4,
        "ytick.major.size": 0,
        "xtick.major.width": AXIS_LINEWIDTH,
        "ytick.major.width": 0,
        "axes.unicode_minus": False,
    })

    matrix = read_matrix(INPUT_XLSX)

    diag = np.diag(matrix)
    row_sum = matrix.sum(axis=1)
    col_sum = matrix.sum(axis=0)

    gross_outflow = row_sum - diag
    gross_inflow = col_sum - diag
    net_change = col_sum - row_sum

    df = pd.DataFrame({
        "class_id": range(1, 11),
        "class_name": CLASS_NAMES,
        "area_2000": row_sum,
        "area_2022": col_sum,
        "diagonal_same_class": diag,
        "gross_outflow": gross_outflow,
        "gross_inflow": gross_inflow,
        "net_change_2022_minus_2000": net_change,
    })
    df["turnover"] = df["gross_inflow"] + df["gross_outflow"]
    df["abs_net_change"] = np.abs(df["net_change_2022_minus_2000"])

    # Use an explicit order so Fig. 1d remains stable and easy to compare with
    # the figure caption.
    df["class_name"] = pd.Categorical(df["class_name"], categories=DISPLAY_ORDER, ordered=True)
    df = df.sort_values("class_name").reset_index(drop=True)
    df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")

    fig = plt.figure(figsize=FIGSIZE, dpi=DPI, facecolor="white")
    ax = fig.add_axes(AX_POS)

    y = np.arange(len(df))
    labels = df["class_name"].astype(str).values
    out_vals = df["gross_outflow"].values.astype(float)
    in_vals = df["gross_inflow"].values.astype(float)

    ax.barh(
        y, -out_vals,
        color=OUTFLOW_COLOR,
        edgecolor="none",
        height=BAR_HEIGHT,
        zorder=3,
        label="Outflow",
    )
    ax.barh(
        y, in_vals,
        color=INFLOW_COLOR,
        edgecolor="none",
        height=BAR_HEIGHT,
        zorder=3,
        label="Inflow",
    )

    ax.invert_yaxis()

    xmax_data = float(max(np.nanmax(out_vals), np.nanmax(in_vals))) if len(df) else 1.0
    # Clean major ticks; the largest observed bar is ~192,585, so 200,000 is readable.
    xmax_tick = 200000.0
    ax.set_xlim(-xmax_tick * (1.0 + LEFT_LABEL_ZONE), xmax_tick * (1.0 + RIGHT_LABEL_ZONE))
    ax.set_ylim(len(df) - 0.35, -0.75)

    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=LABEL_FONT, color=TEXT_COLOR)

    xticks = np.array([-200000, -100000, 0, 100000, 200000], dtype=float)
    ax.set_xticks(xticks)
    ax.set_xticklabels(
        [fmt_full(abs(t)) if abs(t) > 1e-9 else "0" for t in xticks],
        fontsize=TICK_FONT,
        color=AXIS_COLOR,
    )
    ax.set_xlabel("Gross transition area", fontsize=AXIS_LABEL_FONT, color=TEXT_COLOR, labelpad=4)

    ax.grid(axis="x", color=GRID_COLOR, linewidth=GRID_LINEWIDTH, zorder=1)
    ax.grid(axis="y", visible=False)
    ax.set_axisbelow(True)
    ax.axvline(0, color=ZERO_COLOR, linewidth=ZERO_LINEWIDTH, zorder=2)

    # Major numeric labels only. All remain outside the bars.
    pad = xmax_tick * LABEL_PAD_FRAC
    for yi, v in zip(y, out_vals):
        if np.isfinite(v) and v >= LABEL_THRESHOLD:
            ax.text(
                -v - pad, yi, fmt_full(v),
                ha="right", va="center",
                fontsize=VALUE_FONT, color=TEXT_COLOR,
                clip_on=False, zorder=6,
            )
    for yi, v in zip(y, in_vals):
        if np.isfinite(v) and v >= LABEL_THRESHOLD:
            ax.text(
                v + pad, yi, fmt_full(v),
                ha="left", va="center",
                fontsize=VALUE_FONT, color=TEXT_COLOR,
                clip_on=False, zorder=6,
            )

    # Keep left + bottom borders to match the rest of Fig. 1.
    for spine in ["top", "right"]:
        ax.spines[spine].set_visible(False)
    ax.spines["left"].set_visible(True)
    ax.spines["left"].set_color(AXIS_COLOR)
    ax.spines["left"].set_linewidth(AXIS_LINEWIDTH)
    ax.spines["bottom"].set_color(AXIS_COLOR)
    ax.spines["bottom"].set_linewidth(AXIS_LINEWIDTH)

    ax.tick_params(axis="x", colors=AXIS_COLOR, labelsize=TICK_FONT, length=2.4, pad=2)
    ax.tick_params(axis="y", length=0, pad=5)

    legend_handles = [
        Patch(facecolor=OUTFLOW_COLOR, edgecolor="none", label="Outflow"),
        Patch(facecolor=INFLOW_COLOR, edgecolor="none", label="Inflow"),
    ]
    # Legend is placed below the chart and centred under the main plotting axis,
    # rather than under the full canvas. This aligns it with the x-axis title
    # and with Fig. 1e during side-by-side assembly.
    fig.legend(
        handles=legend_handles,
        loc="lower center",
        bbox_to_anchor=(LEGEND_X, LEGEND_Y),
        ncol=2,
        frameon=False,
        fontsize=LEGEND_FONT,
        handlelength=1.12,
        columnspacing=0.70,
        borderaxespad=0,
    )

    # Fixed-canvas export for downstream assembly. Do not use bbox_inches='tight'.
    if EXPORT_TRANSPARENT:
        fig.patch.set_alpha(0)
        for a in fig.axes:
            a.set_facecolor("none")

    save_kwargs = dict(transparent=EXPORT_TRANSPARENT)
    fig.savefig(OUTPUT_PNG, dpi=DPI, **save_kwargs)
    fig.savefig(OUTPUT_PDF, **save_kwargs)
    fig.savefig(OUTPUT_SVG, **save_kwargs)
    plt.close(fig)

    print(f"Saved: {OUTPUT_PNG}")
    print(f"Saved: {OUTPUT_PDF}")
    print(f"Saved: {OUTPUT_SVG}")
    print(f"Saved: {OUTPUT_CSV}")


if __name__ == "__main__":
    main()
